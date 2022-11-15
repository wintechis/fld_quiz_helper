from openpyxl import load_workbook, worksheet
from typing import List, Union


def xl_to_lst(wks: worksheet, last_col: str) -> List[Union[str, int]]:
    tpl_header = tuple(map(lambda cell: cell.value, wks['A1':f'{last_col}1'][0]))
    lst = []
    for row in range(2, 500):
        if wks.cell(row, 1).value == None: break 
        dct = {}
        for col in range(1,  len(tpl_header) + 1):
            dct[tpl_header[col-1]] = wks.cell(row,col).value
        lst.append(dct)
    return lst
        


workbook = load_workbook(filename="data.xlsx")

dct_data = {}
for (name, last_col) in [('chapters', 'B'), ('learningGoals', 'C'), ('quizItems', 'F')]:
    if name in workbook.sheetnames:
        dct_data[name] = xl_to_lst(workbook[name], last_col)
        
dct_chapters = {}
